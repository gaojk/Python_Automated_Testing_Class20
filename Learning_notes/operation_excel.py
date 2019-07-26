"""
操作EXCEL
"""
from openpyxl import load_workbook

# wb = load_workbook("testcase.xlsx")  # 创建workbook对象
# ws = wb.active  # 打开第一个sheet，创建sheet对象
# # ws = wb["test"]    # 打开指定的sheet页
# # wb.create_sheet(title="test", index=1)
# print(ws.max_row, ws.max_column)
#
# cell = ws.cell(row=1, column=1)    # 定位cell单元格，创建cell对象(代表第一行第一列的单元格)
# cell.value  # 获取cell对象的值
# print(cell.value)
# print(ws["A1"])
# print(ws[1], ws["A"])
# print(cell.value)
# ws.iter_rows(min_row=1, max_row=2, min_col=2, max_col=4, values_only=True)

wb = load_workbook("testcase.xlsx")  # 创建workbook对象
ws = wb["test"]    # 打开指定的sheet页
ws.cell(row=1, column=6, value="第一行第六列写入值")
ws.cell(row=1, column=7, value="第一行第七列写入值")
wb.save("testcase.xlsx")
wb.close()
