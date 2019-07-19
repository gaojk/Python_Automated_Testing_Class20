from openpyxl import load_workbook

wb = load_workbook("case_data.xlsx")    # 创建excel对象
ws = wb.active    # 创建表单对象
# ws = wb["divide"]

one_cell = ws.cell(row=2, column=3)    # 获取单元格值
# 更新单元格的值的两种方法
# one_cell.value = 3
# ws.cell(2, 3, value="haha")

# 保存excel文件
# wb.save("case_data.xlsx")

# for row in range(ws.min_row, ws.max_row+1):
# 获取第一行的值
one_low = []
for col in range(ws.min_column, ws.max_column+1):
    one_low.append(ws.cell(1, col).value)

# 获取表头值和数据值，将他们分别转为元组类型再通过zip函数转换为字典做到表头和数据值一一对应
# first_line_data = tuple(ws.iter_rows(max_row=1, values_only=True))  # 通过iter_rows方法获取第一行表头的值，并转换成元组
# other_line_data = tuple(ws.iter_rows(min_row=2, values_only=True))  # 获取其他行的测试数据
# test_data = []
# for i in other_line_data:
#     data_dict = dict(zip(first_line_data[0], i))
#     test_data.append(data_dict)
# print(test_data)
# 获取其他行的值
data_list = []
for row in range(ws.min_row+1, ws.max_row+1):
    one_list = []
    for col in range(ws.min_column, ws.max_column + 1):
        one_list.append(ws.cell(row, col).value)
    data_list.append(one_list)

# 将第一行的值和其他行的值组成字典
def read_data():
    testdata = []
    for i in range(len(data_list)):
        data_dict = dict(zip(one_low, data_list[i]))
        testdata.append(data_dict)
    return testdata


if __name__ == '__main__':
    print(read_data())

